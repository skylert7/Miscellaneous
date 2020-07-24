import requests
import urllib3
import whois
import pywhois
import urllib2
from openpyxl import load_workbook
from openpyxl import Workbook
headers = {'content-type': 'application/json', 'Accept-Charset': 'UTF-8'}

book = load_workbook('Data.xlsx')
sheetRead = book.active
# a1 = sheet['A1']
# a2 = sheet['A2']
# a3 = sheet.cell(row=3, column=1)
# print a1.value
# print a2.value
# print a3.value

result = ''

bookWrite = Workbook()
sheetWrite = bookWrite.active
for i in xrange(381, 386):
    aCol = sheetRead['A' + str(i)]
    # bCol = sheetRead['B' + str(i)]
    # sheetWrite['B' + str(i)] = bCol.value
    url = 'http://' + aCol.value
    try:
        data = urllib2.urlopen(url)
        if data.getcode() == 200:
            sheetWrite['C' + str(i)] = data.geturl()
    except urllib2.HTTPError, e:
        result = 'None'
        sheetWrite['C' + str(i)] = result
    except urllib2.URLError, e:
        result = 'None'
        sheetWrite['C' + str(i)] = result

bookWrite.save("381to385.xlsx")

